# to Exe file  "pyinstaller -wF main.py"
from ast import Break, For, If
import re
from site import venv
from xml.dom.minidom import CharacterData
from xml.etree.ElementTree import tostring
import PySimpleGUI as sg
import datetime , os
import openpyxl as xl_read
import xlsxwriter as xl_writer
import json
import subprocess

e = datetime.datetime.now()
time = (" %s/%s/%s " % (e.day, e.month, e.year )) #%s:%s:%s , e.hour, e.minute, e.second
time_day = e.day

# current_month = e.month
# _time = ("%s-%s-%s" % ( e.year, e.month,e.day )) #%s:%s:%s , e.hour, e.minute, e.second
# time =  datetime.fromisoformat(_time)
# time = datetime.date.fromisoformat('2019-12-04')

path = "C:\\Users\\FelixLUI\\workload_function\\Tasks_Felix.xlsx"
JsonPath = "C:\\Users\\FelixLUI\\workload_function\\Description_json.json"
wb_obj = xl_read.load_workbook(path)
workbook = xl_writer.Workbook('Tasks_Felix.xlsx')
text_version = "Created by FelixLui v2.1"

# systemtype = {'Salesforce', 'CMS', 'LIMS', 'CSS'}
systemtype = []
SYSTEMTYPE_itme = []
SYSTEMTYPE_id = 0
Descriptiontype = []
defaultName = "Felix"
ProblemType = []
# Person_In_Charge = [ "YiZhan", "Danny", "Tao"]
Person_In_Charge = []
ProblemCategory = []
# ProblemType = []
ProblemType_selete = None
# ProblemCategory_list = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
# Record_max = [10,20,50]
Record_max = []
Start_Q_Data = ""

####              Salesforce CMS LIMS CSS OA Extra    ###
# Start_Q_index = [79, 20, 12, 40, 19, 70] # start with the date before
Start_Q_index = [] # please enter the start data -1

max_row = 0  # 直

baseRowCount = 2

# showMaxRow = 10 # not use, but 
# SF_Descriptiontype = []
# CMS_Descriptiontype = []
# CSS_Descriptiontype = []
# LIMS_Descriptiontype = []

ErrorMessage = 'Please DONOT Open the excel and the Program in the same time'

json_data = [] # your list with json objects (dicts)

######   initialization  data from Json
with open(JsonPath , encoding='utf-8') as json_file:
    json_data = json.load(json_file)
    # print (json_data)
    json_data_sys = json_data["SystemtypeList"]
    for item in json_data_sys:   # add systemtype with Json Key
        systemtype.append(item)
        # print(item)
        

    json_data_Start_Q_index = json_data["Start_Q_index"]
    for item in json_data_Start_Q_index:   # add systemtype with Json Key
        Start_Q_index.append(item)

    json_data_Record_max = json_data["Record_max"]
    for item in json_data_Record_max:   # add systemtype with Json Key
        Record_max.append(item)

        
    json_data_Person_In_Charge= json_data["Person_In_Charge"]
    for item in json_data_Person_In_Charge:   # add systemtype with Json Key
        Person_In_Charge.append(item)

    json_data_ProblemCategory= json_data["ProblemCategory"]
    for item in json_data_ProblemCategory:   # add systemtype with Json Key
        ProblemCategory.append(item)

    
    json_data_ProblemType= json_data["ProblemType"]
    for item in json_data_ProblemType:   # add systemtype with Json Key
        ProblemType.append(item)


    Start_Q_Data = json_data["Start_Q_Data"]
    max_row = json_data["Max_row"]
    

sg.theme('Dark Brown')    

layout_1 = [

    [sg.Text("Auto fill-in Excel" , justification='center',size=(50,1))],
    
    [sg.HorizontalSeparator()]] + [
    [sg.Checkbox(f'{Person_In_Charge[i]}', enable_events=True, key=f'{Person_In_Charge[i]}')
        for i in  range(len(Person_In_Charge)) ],
    # [sg.InputText(time_day)],
    [sg.Text(time), sg.InputText(time_day, enable_events=True,size=(2, 2), key='-TIMEDAY-' ),  sg.Text( "Start Q:" + Start_Q_Data + str(Start_Q_index) )],
    # [sg.Text('Checkbox'), sg.Checkbox('Felix',default=True , key="-FELIX-") ,sg.Checkbox('Yi Zhan' , key="-YIZHAN-"),sg.Checkbox('Danny' , key="-DANNY-"),sg.Checkbox('Tao' , key="-TAO-") ],
    

    [sg.Text('Name:'), sg.InputText('', key='-NAME-')],
    [sg.Text('SystemType:'), sg.Combo(list(systemtype), key='-SYSTEMTYPE-',enable_events=True, size=(15, 6)) , sg.Text('Record max:'), sg.Combo(list(Record_max), key='-RECORDMAX-',default_value= Record_max[0], size=(10, len(Record_max)))  ],
    [sg.Text('Description:'), sg.Combo([], key='-DESCRIPTION-',enable_events=True,size=(40, 10))],
    [sg.Text('Problem Type:'), sg.Combo(list(ProblemType), key='-PROBLEMTYPE-',enable_events=True,size=(40, 10))],
    [sg.Text('Remark:'), sg.Input('', key='-REMARK-')],
    [sg.Button('OK' , size=(50, 2))],
    [sg.Button('Open Excel') ,sg.Button('Open Json') , sg.Button('Analysis update'), sg.Button('Testing'),sg.Button('Cancel')],
    [sg.Button('Open File location')],
    [sg.Text( text_version , justification='center',size=(50,1))],
    [sg.Text( ErrorMessage , justification='center',size=(50,1) , text_color = 'Red', background_color= 'Yellow')],
    
    # [sg.HorizontalSeparator()]] + [
    # [sg.Checkbox(f'{Person_In_Charge[i]}', enable_events=True, key=f'{Person_In_Charge[i]}')
    #     for i in  range(len(Person_In_Charge)) ] 
    # [[sg.Checkbox(f'{Person_In_Charge[i]}', enable_events=True, key=f'{Person_In_Charge[i]}')
    #     for i in  range(len(Person_In_Charge)) ] ]

]
# For now will only show the name of the file that was chosen
layout_2 = [[sg.MLine(key='-ML1-'+sg.WRITE_ONLY_KEY,  size=(75,20))]]
# ----- Full layout -----
layout = [
    [sg.Column(layout_1),
     sg.VSeperator(),
     sg.Column(layout_2),]
]
window = sg.Window("Column Demo", layout)

def Analysis_update_func():
    systemtype_lenght = 1
    systemtype_Q = 0

    problem_C_list = [0] * len(ProblemCategory)

    for page in systemtype: 
        counting_row = 0
        
        # problem_C_list = [0] * len(ProblemCategory)

        systemtype_lenght += 1

        sheet_obj = wb_obj[page]
        # sheet_obj = wb_obj["Salesforce"]
        # max_row = sheet_obj.max_row  # 直
        # max_row = 1000  # 直


        #####################################   base add Analysis 
        a_test = (baseRowCount + Start_Q_index[systemtype_Q], max_row + 1)
        print(Start_Q_index[systemtype_Q],  sheet_obj , a_test )

        for i in range(baseRowCount + Start_Q_index[systemtype_Q], max_row + 1):
            cell_obj = sheet_obj.cell(i, 1)

            problemType_cell_obj = sheet_obj.cell(i, 7)
            # print(i)
            if cell_obj.value != None:
                counting_row += 1


                #########  problem counting
                for k in range( len(ProblemCategory) ) :
                    if problemType_cell_obj.value == ProblemCategory[k]:
                        # print(problemType_cell_obj.value)
                        problem_C_list[k] += 1
                        


            else:
                # print("counting_row: "+ str(counting_row) ) # Row
                # print(problem_C_list)
                # print(systemtype_lenght , counting_row ,problem_C_list )
                break
            # print(problem_C_list )
            
        systemtype_Q += 1
        sheet_obj_Analysis = wb_obj["Analysis"]

        # print(counting_row)
        sheet_obj_Analysis.cell(systemtype_lenght,2).value = counting_row

        for i in range(baseRowCount , len(ProblemCategory)):
            sheet_obj_Analysis.cell( baseRowCount + i ,6, problem_C_list[i])
        
        wb_obj.save(path)
        #######################################



# window['-SYSTEMTYPE-'].update(value= "Salesforce")
# Create an event loop
while True:
    event, values = window.read()

    

    if event == 'Testing':

        print(len(ProblemCategory))



        # # def Analysis_update_func():
        # systemtype_lenght = 1
        # systemtype_Q = 0
        # for page in systemtype: 
        #     counting_row = 0

        #     systemtype_lenght += 1

        #     sheet_obj = wb_obj[page]
        #     # sheet_obj = wb_obj["Salesforce"]
        #     # max_row = sheet_obj.max_row  # 直
        #     # max_row = 1000  # 直


        #     #####################################   base add Analysis 
        #     print(Start_Q_index[systemtype_Q])
        #     for i in range(baseRowCount + Start_Q_index[systemtype_Q], max_row + 1):

        #         cell_obj = sheet_obj.cell(i, 1)
        #         Category_obj = sheet_obj.cell(i, 7)
        #         Category_obj_value = Category_obj.value

        #         # print(Category_obj_value)
        #         if cell_obj.value != None:
        #             counting_row += 1
        #             for j in range(len(ProblemCategory)):
        #                 # print(Category_obj_value)
        #                 if Category_obj_value == ProblemCategory[j]:
        #                     print(Category_obj_value)
        #                 # else:
                        

                        
        #         else:
        #             # print("counting_row: "+ str(counting_row) ) # Row
        #             # print(counting_row)
        #             print(systemtype_lenght , counting_row)
        #             break
                
                
        #     systemtype_Q += 1
        #     sheet_obj_Analysis = wb_obj["Analysis"]
        #     # sheet_obj_Analysis.cell(2,systemtype_lenght).value = counting_row
            
        #     sheet_obj_Analysis.cell(systemtype_lenght,2, counting_row)
            
        #     wb_obj.save(path)
        #     #######################################





    if event == sg.WIN_CLOSED or event == 'Cancel': # if user closes window or clicks cancel
        break
    if event == 'Open Excel':
        # f = open(path , "r")
        os.startfile(path)
        # break
    if event == 'Open Json':
        os.startfile(JsonPath)
        # break
    
    if event == 'Analysis update':
        # os.startfile(JsonPath)
        Analysis_update_func()
    
    # if event == 'Close Popup':
    #     window.close()
    #     break
    
    if event == 'Open File location':
        # os.startfile(JsonPath)
        subprocess.Popen(r'explorer /select,"C:\Users\FelixLUI\workload_function"')

    if event == '-SYSTEMTYPE-':
        item = values[event]

        Descriptiontype.clear()     # Clear all Description data
        # ProblemType.clear() 

        sheet_obj = wb_obj[values['-SYSTEMTYPE-']]
        # print(json_data[str(item)] )
        for item in json_data_sys[str(item)]:     # add Description list json
            # print( json.dumps(list(item.values())[0] ) )
            # print( item)
            SYSTEMTYPE_itme.append(item)
            # Descriptiontype.append(item)
            Descriptiontype.append(item["name"])
            # ProblemType.append(item["id"])

            window['-DESCRIPTION-'].update(values = Descriptiontype)

        
        sheet_obj = wb_obj[values['-SYSTEMTYPE-']]
        # max_row = sheet_obj.max_row  # 直
        # max_row = 1000  # 直

    

                

##################################################################################
############################## text box with excel data function   ##########################################
##################################################################################
 
        # max_row_i = max_row
        
        # window['-ML1-'+sg.WRITE_ONLY_KEY].print("↓↓↓↓↓↓↓↓↓↓ " + values['-SYSTEMTYPE-'] , text_color='red', background_color='yellow')
             
        # for i in range(3, max_row):
        #     cell_obj_row = sheet_obj.cell(i, 1)

        #     record_row = []  
              
        #     if cell_obj_row.value != None:
        #         for j in range(1,7):
        #             # print(i , j , max_row_i)
        #             cell_obj_row_column = sheet_obj.cell(i, j)

        #             record_row.append(cell_obj_row_column.value )

        #             # if j == 2: # is mounth XX/X/XXXX
        #             #     # print(cell_obj_row_column.value) 
        #             #     for k in  range(0, len( str(cell_obj_row_column.value) )):
        #             #         # print( str(cell_obj_row_column.value) [k])
        #             #         _character = str(cell_obj_row_column.value)[k]
        #             #         if _character == "/":
        #             #             # print (str(cell_obj_row_column.value)[k+1])
        #             #             counted_month =  int(str(cell_obj_row_column.value)[k+1] )
        #             #             print(current_month, counted_month)
        #             #             if current_month == counted_month:
        #             #                 for o in range(1,7):
        #             #                     cell_obj_row_column_ = sheet_obj.cell(i, o)
        #             #                     record_row.append(cell_obj_row_column_.value )

                                
        #         window['-ML1-'+sg.WRITE_ONLY_KEY].print(record_row)
        #     else:
        #         window['-ML1-'+sg.WRITE_ONLY_KEY].print("↑↑↑↑↑↑↑↑↑↑ " + values['-SYSTEMTYPE-'] , text_color='red', background_color='yellow')
        #         break




        showMaxRow = values['-RECORDMAX-']

        max_row_i = 0
        
        window['-ML1-'+sg.WRITE_ONLY_KEY].print("↓↓↓↓↓↓↓↓↓↓ " + values['-SYSTEMTYPE-'] + "      Only show "+str(showMaxRow)+" record "  , text_color='red', background_color='yellow')
             
        for i in range(baseRowCount, max_row):
            cell_obj_row = sheet_obj.cell(i, 1)
 

            if cell_obj_row.value == None:
                break
            max_row_i += 1
        # print(max_row_i)
        
        startMaxRow = (  max_row_i - showMaxRow)
        _startMaxRow = 0
        if startMaxRow > 0:
            _startMaxRow = startMaxRow
            max_row_i = showMaxRow     # reset the max row if over 50 set to 50 

        for i in range(max_row_i):  # max_row_i: nounal count
            _i = (i + baseRowCount )              #base start row
            _i += _startMaxRow      #add the i with showMaxRow
            cell_obj_row = sheet_obj.cell( _i , 1)
            record_row = [] 
            for j in range(1,7):
                cell_obj_row_column = sheet_obj.cell( _i, j)
                record_row.append(cell_obj_row_column.value )

                # print(record_row)
            window['-ML1-'+sg.WRITE_ONLY_KEY].print(record_row)
        #     else:
        window['-ML1-'+sg.WRITE_ONLY_KEY].print("↑↑↑↑↑↑↑↑↑↑ " + values['-SYSTEMTYPE-']+ "      Only show "+str(showMaxRow)+" record " , text_color='red', background_color='yellow')
        #         break


                # for j in range(1,7):
                    # print(i , j , max_row_i)
                    # cell_obj_row_column = sheet_obj.cell(i, j)

                    # record_row.append(cell_obj_row_column.value )

                                
                # window['-ML1-'+sg.WRITE_ONLY_KEY].print(record_row)
            # else:
                # window['-ML1-'+sg.WRITE_ONLY_KEY].print("↑↑↑↑↑↑↑↑↑↑ " + values['-SYSTEMTYPE-'] , text_color='red', background_color='yellow')
                # break

##################################################################################
############################## text box with excel data function   END  ##########################################
##################################################################################

            # record_row.clear()
        # window['-OUTPUT-'].update(record_list)
        # print(record_list)
        # record_list.clear()

        # if str(item) == 'Salesforce':
        #     print("==Testing")
        #     window['-DESCRIPTION-'].update(values = SF_Descriptiontype) 
        # elif  str(item) == 'CMS':
        #    window['-DESCRIPTION-'].update(values = CMS_Descriptiontype) 
        # elif  str(item) == 'LIMS':
        #     window['-DESCRIPTION-'].update(values = LIMS_Descriptiontype) 
        # elif  str(item) == 'CSS':
        #     window['-DESCRIPTION-'].update(values = CSS_Descriptiontype) 
    if event == '-DESCRIPTION-':
        item = values[event]

        print(SYSTEMTYPE_itme)
        print("  in  -DESCRIPTION- " )
        # for item_ in json_data_sys[str(SYSTEMTYPE_itme)]:     # add Description list json
            # print( json.dumps(list(item.values())[0] ) )
            # if item == item_:

        window['-PROBLEMTYPE-'].update(values = ProblemType)   # clean ProblemTypeBar
        ProblemType_selete = None

        # if item == None:

        for i in range(len(SYSTEMTYPE_itme)):
            
            if item == SYSTEMTYPE_itme[i]["name"]:
                # print(SYSTEMTYPE_itme[i]["id"])
                SYSTEMTYPE_id = SYSTEMTYPE_itme[i]["id"]

                window['-PROBLEMTYPE-'].update(value= SYSTEMTYPE_id)
        # else:
        #     SYSTEMTYPE_id = item
    
    if event == '-PROBLEMTYPE-':
        item = values[event]

        ProblemType_selete = item 
        # print(item)
        


    if event == 'OK':
            
        print('SYSTEMTYPE: ', values['-SYSTEMTYPE-']  ) #values['-SYSTEMTYPE-']
        
        if values['-SYSTEMTYPE-'] == "":
            sg.Popup( " ERROR: out of SYSTEMTYPE" )
            break

        # sheet_obj = wb_obj.active
        sheet_obj = wb_obj[values['-SYSTEMTYPE-']]
        # max_row = sheet_obj.max_row  # 直
        # max_row = 1000  # 直

        # print( type(sheet_obj) )
        # if sheet_obj == "<class 'openpyxl.worksheet.worksheet.Worksheet'>":
        #     sg.Popup( 'Please enter system type' )

        ## time update
        time_day = values['-TIMEDAY-']
        time_new = (" %s/%s/%s " % (time_day, e.month, e.year )) 
            
        sg.Popup( values['-NAME-'], values['-DESCRIPTION-'], values['-REMARK-'] , time_new ) # Popup(event, values, values['-NAME-'])
        # Popup_layout = [[sg.MLine(key='-ML1-'+sg.WRITE_ONLY_KEY,  size=(60,10))]]


        for i in range(baseRowCount, max_row + 1):
            # cell_obj = sheet_obj.cell(1, i)
            # print(cell_obj.value)
            cell_obj = sheet_obj.cell(i, 1)

            if i == max_row:
                ### Error
                sg.Popup( "Failed add Row, ERROR: out of max_row" +str(max_row) )
                break

            if cell_obj.value != None:
                print(cell_obj.value,  i ,  max_row)

            else:
                print(" in == None") # Row
                # sheet_obj['F6'] = 'Writing new Value!'
                # print( sheet_obj['F6'].value) # Row
                cell_number = 'cell_obj.value'
                
                # add row
                sheet_obj.cell( i , 1).value =  (i -2)
                sheet_obj.cell( i , 2).value =  time_new
                sheet_obj.cell( i , 3).value =   values['-DESCRIPTION-']
                sheet_obj.cell( i , 4).value =   values['-NAME-']

                nameText = ""
                for j in  range(len(Person_In_Charge)):
                    if values[f'{Person_In_Charge[j]}'] == True:
                        nameText += " "+ Person_In_Charge[j] 
                print(nameText)
                sheet_obj.cell( i , 5).value =   defaultName + nameText
                    
                
                # sheet_obj.cell( i , 5).value =   "Felix"
                sheet_obj.cell( i , 6).value =   values['-REMARK-']
                # sheet_obj['A32'] = 1

                if ProblemType_selete != None:
                    sheet_obj.cell( i , 7).value = ProblemType_selete
                else:
                    sheet_obj.cell( i , 7).value = SYSTEMTYPE_id
                # mycell = sheet_obj.cell(30, 1)
                # mycell ='Writing data to E4'
                # wb_obj.save()
                # wb_obj.save('example_filetest.xlsx')
                wb_obj.save(path)

                # 'Record added: ' +  i,

                break
                # cell_obj.value = cell_obj.value+1
        
    

window.close()

